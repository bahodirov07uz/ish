# crm/views/excel.py
"""
Excel orqali ish qo'shish — 3 endpoint:
  GET  /excel/shablon/<ishchi_id>/       → .xlsx fayl yuklab olish
  POST /excel/parse/<ishchi_id>/         → parse + validatsiya + preview JSON
  POST /excel/saqlash/<ishchi_id>/       → tasdiqlangan satrlarni saqlash
"""
import io
import json
import logging
from datetime import datetime
from decimal import Decimal, InvalidOperation

from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse, JsonResponse,FileResponse
from django.shortcuts import get_object_or_404,render
from django.utils import timezone
from django.db import transaction
from django.views.decorators.http import require_http_methods
import urllib.parse

import openpyxl

import crm.models as m
from xomashyo.models import Xomashyo, XomashyoVariant, XomashyoCategory

logger = logging.getLogger(__name__)


def is_admin(user):
    return user.is_staff or user.is_superuser


@login_required(login_url='login')
@user_passes_test(is_admin, login_url='login')
def excel_ish_sahifa(request):
    context = {
        'ishchilar': m.Ishchi.objects.filter(
            is_oylik_open=True
        ).select_related('turi').order_by('ism'),
    }
    return render(request, 'excel/excel_ish_qoshish.html', context)

# ══════════════════════════════════════════════════════════════════
# 1.  SHABLON YUKLAB OLISH
# ══════════════════════════════════════════════════════════════════
@login_required(login_url='login')
@user_passes_test(is_admin, login_url='login')
def excel_shablon(request, ishchi_id):
    """GET — ishchi turiga mos .xlsx shablon qaytaradi"""
    from .excel_shablon_maker import get_shablon_bytes   # pastdagi make_templates.py

    ishchi = get_object_or_404(m.Ishchi, pk=ishchi_id)
    turi   = (ishchi.turi.nomi if ishchi.turi else 'pardoz').lower()

    # Ma'lumotnoma listalari — dropdown uchun
    mahsulotlar = list(m.Product.objects.values_list('nomi', flat=True).order_by('nomi'))

    terilar = list(
        Xomashyo.objects.filter(
            category__name__iexact='teri',
            category__turi='real',
            holati='active',
            miqdori__gt=0
        ).values_list('nomi', flat=True)
    )

    astarlar = list(
        Xomashyo.objects.filter(
            category__name__iexact='astar',
            category__turi='real',
            holati='active',
            miqdori__gt=0
        ).values_list('nomi', flat=True)
    )

    padojlar = list(
        Xomashyo.objects.filter(
            category__name__iexact='padoj',
            category__turi='real',
            holati='active',
        ).values_list('nomi', flat=True)
    )

    kroy_zak = list(
        Xomashyo.objects.filter(
            category__name__iexact='kroy',
            category__turi='process',
            holati='active',
            miqdori__gt=0
        ).values_list('nomi', flat=True)
    )

    zakatovka_list = list(
        Xomashyo.objects.filter(
            category__name__iexact='zakatovka',
            category__turi='process',
            holati='active',
            miqdori__gt=0
        ).values_list('nomi', flat=True)
    )

    kwargs = dict(
        ishchi_ism  = f"{ishchi.ism} {ishchi.familiya}",
        mahsulotlar = mahsulotlar,
    )
    if turi in ('kroy', 'rezak'):
        kwargs.update(terilar=terilar, astarlar=astarlar)
    elif turi == 'zakatovka':
        kwargs.update(kroy_xomashyolar=kroy_zak)
    elif turi == 'kosib':
        kwargs.update(padojlar=padojlar, zakatovkalar=zakatovka_list)

    content = get_shablon_bytes(turi, **kwargs)
    fname_safe = f"shablon_{turi}.xlsx"
    fname_full = urllib.parse.quote(f"shablon_{turi}_{ishchi.ism}.xlsx")

    buf = io.BytesIO(content)
    resp = FileResponse(
        buf,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        filename=fname_safe,
    )
    resp['Content-Disposition'] = (
        f"attachment; filename=\"{fname_safe}\"; "
        f"filename*=UTF-8''{fname_full}"
    )
    return resp



# ══════════════════════════════════════════════════════════════════
# 2.  PARSE + VALIDATSIYA (preview JSON)
# ══════════════════════════════════════════════════════════════════
@login_required(login_url='login')
@user_passes_test(is_admin, login_url='login')
@require_http_methods(['POST'])
def excel_parse(request, ishchi_id):
    """
    POST multipart/form-data  file=<xlsx>
    → JSON { satrlar: [...], xatolar_soni, ogohlantirish_soni, ok_soni }
    """
    ishchi = get_object_or_404(m.Ishchi, pk=ishchi_id)
    turi   = (ishchi.turi.nomi if ishchi.turi else 'pardoz').lower()

    if 'file' not in request.FILES:
        return JsonResponse({'error': 'Fayl yuklanmadi'}, status=400)

    xlsx_file = request.FILES['file']
    try:
        wb = openpyxl.load_workbook(xlsx_file, data_only=True)
    except Exception as e:
        return JsonResponse({'error': f'Excel o\'qib bo\'lmadi: {e}'}, status=400)

    ws = wb.active

    # Ustun sarlavhalarini topish (4-satr = header, 1-satr title)
    header_row = _find_header_row(ws)
    if header_row is None:
        return JsonResponse({'error': 'Sarlavha satri topilmadi (4-satr tekshirildi)'}, status=400)

    col_map = _build_col_map(ws, header_row)

    satrlar = []
    for row_idx in range(header_row + 2, ws.max_row + 1):   # +2: header + namuna
        row_vals = {
            key: _cell_val(ws.cell(row_idx, col))
            for key, col in col_map.items()
        }
        # Bo'sh satrlarni o'tkazib yuborish
        if not any(v for v in row_vals.values()):
            continue

        result = _validate_row(row_idx - header_row - 1, row_vals, ishchi, turi)
        satrlar.append(result)

    ok_soni           = sum(1 for s in satrlar if s['holat'] == 'ok')
    ogohlantirish_soni = sum(1 for s in satrlar if s['holat'] == 'ogohlantirish')
    xatolar_soni      = sum(1 for s in satrlar if s['holat'] == 'xato')

    return JsonResponse({
        'satrlar'           : satrlar,
        'ok_soni'           : ok_soni,
        'ogohlantirish_soni': ogohlantirish_soni,
        'xatolar_soni'      : xatolar_soni,
        'jami'              : len(satrlar),
        'ishchi_id'         : ishchi_id,
        'turi'              : turi,
    })


# ══════════════════════════════════════════════════════════════════
# 3.  SAQLASH (tasdiqlangan satrlar)
# ══════════════════════════════════════════════════════════════════
@login_required(login_url='login')
@user_passes_test(is_admin, login_url='login')
@require_http_methods(['POST'])
def excel_saqlash(request, ishchi_id):
    """
    POST JSON body:
    {
      "satrlar": [...],          ← parse dan kelgan satrlar (holat != 'xato')
      "xomashyo_map": {           ← foydalanuvchi biriktirgan xomashyo ID lari
        "mahsulot_nomi:teri": xomashyo_id,
        ...
      },
      "skip_xatolar": true,
      "include_ogohlantirishlar": true
    }
    → JSON { yaratilgan: N, o'tkazilgan: N }
    """
    ishchi = get_object_or_404(m.Ishchi, pk=ishchi_id)
    turi   = (ishchi.turi.nomi if ishchi.turi else 'pardoz').lower()

    try:
        body = json.loads(request.body)
    except Exception:
        return JsonResponse({'error': 'JSON formati noto\'g\'ri'}, status=400)

    satrlar                  = body.get('satrlar', [])
    xomashyo_map             = body.get('xomashyo_map', {})
    skip_xatolar             = body.get('skip_xatolar', True)
    include_ogohlantirishlar = body.get('include_ogohlantirishlar', True)

    # Filter
    to_save = []
    for s in satrlar:
        if s['holat'] == 'xato' and skip_xatolar:
            continue
        if s['holat'] == 'ogohlantirish' and not include_ogohlantirishlar:
            continue
        to_save.append(s)

    yaratilgan  = 0
    otkazilgan  = 0
    xatolar     = []

    with transaction.atomic():
        for satr in to_save:
            try:
                _satr_saqlash(satr, ishchi, turi, xomashyo_map)
                yaratilgan += 1
            except Exception as e:
                logger.exception(f"Satr #{satr.get('satr_num')} saqlanmadi: {e}")
                xatolar.append({'satr_num': satr.get('satr_num'), 'xato': str(e)})
                otkazilgan += 1

    return JsonResponse({
        'yaratilgan': yaratilgan,
        'otkazilgan': otkazilgan,
        'xatolar'   : xatolar,
    })


# ══════════════════════════════════════════════════════════════════
# YORDAMCHI FUNKSIYALAR
# ══════════════════════════════════════════════════════════════════

def _find_header_row(ws):
    """Sarlavha satrini topish — 'mahsulot' so'zi bor satr"""
    for row_idx in range(1, min(10, ws.max_row + 1)):
        for cell in ws[row_idx]:
            v = str(cell.value or '').lower()
            if 'mahsulot' in v:
                return row_idx
    return None


def _build_col_map(ws, header_row):
    """
    Header satridan ustun nomlarini → col index mapping
    Qaytadi: {'mahsulot_nomi': 1, 'soni': 2, ...}
    """
    KEY_ALIASES = {
        'mahsulot': 'mahsulot_nomi',
        'mahsulot_nomi': 'mahsulot_nomi',
        'soni': 'soni',
        'ish_sanasi': 'ish_sanasi',
        'ish sanasi': 'ish_sanasi',
        'sana': 'ish_sanasi',
        'teri_nomi': 'teri_nomi',
        'teri nomi': 'teri_nomi',
        'teri': 'teri_nomi',
        'teri_variant_rang': 'teri_variant_rang',
        'teri variant rang': 'teri_variant_rang',
        'teri_sarfi': 'teri_sarfi',
        'teri sarfi': 'teri_sarfi',
        'astar_nomi': 'astar_nomi',
        'astar nomi': 'astar_nomi',
        'astar': 'astar_nomi',
        'astar_variant_rang': 'astar_variant_rang',
        'astar variant rang': 'astar_variant_rang',
        'astar_sarfi': 'astar_sarfi',
        'astar sarfi': 'astar_sarfi',
        'kroy_xomashyo': 'kroy_xomashyo',
        'kroy xomashyo': 'kroy_xomashyo',
        'kroy': 'kroy_xomashyo',
        'mustaqil_ish': 'mustaqil_ish',
        'mustaqil ish': 'mustaqil_ish',
        'padoj_nomi': 'padoj_nomi',
        'padoj nomi': 'padoj_nomi',
        'padoj': 'padoj_nomi',
        'padoj_variant_rang': 'padoj_variant_rang',
        'padoj variant rang': 'padoj_variant_rang',
        'variant_rang': 'variant_rang',
        'variant rang': 'variant_rang',
        'variant_razmer': 'variant_razmer',
        'variant razmer': 'variant_razmer',
        'zakatovka_nomi': 'zakatovka_nomi',
        'zakatovka nomi': 'zakatovka_nomi',
        'zakatovka': 'zakatovka_nomi',
        'izoh': 'izoh',
    }
    col_map = {}
    for cell in ws[header_row]:
        raw = str(cell.value or '').strip()
        # Muhim qismni ajratib olish (masalan "Mahsulot nomi *" → "mahsulot nomi")
        cleaned = raw.lower().replace('*', '').replace('(', '').replace(')', '').strip()
        # Birinchi 3 so'z bilan ham solishtiramiz
        short = ' '.join(cleaned.split()[:3])
        key   = KEY_ALIASES.get(cleaned) or KEY_ALIASES.get(short)
        if key:
            col_map[key] = cell.column
    return col_map


def _cell_val(cell):
    """Katak qiymatini string sifatida qaytarish"""
    v = cell.value
    if v is None:
        return ''
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    return str(v).strip()


def _validate_row(satr_num, vals, ishchi, turi):
    """
    Bir satrni validatsiya qilish.
    Qaytadi: {satr_num, holat, xatolar, ogohlantirishlar, ...vals}
    """
    xatolar         = []
    ogohlantirishlar = []
    computed        = {}

    # ── Majburiy: mahsulot ──────────────────────────────────────
    mahsulot_nomi = vals.get('mahsulot_nomi', '').strip()
    if not mahsulot_nomi:
        xatolar.append('mahsulot_nomi bo\'sh')
    else:
        mahsulot = m.Product.objects.filter(nomi__iexact=mahsulot_nomi).first()
        if not mahsulot:
            xatolar.append(f'Mahsulot topilmadi: "{mahsulot_nomi}"')
        else:
            computed['mahsulot_id'] = mahsulot.id

    # ── Majburiy: soni ──────────────────────────────────────────
    try:
        soni = int(float(vals.get('soni', 0)))
        if soni < 1:
            raise ValueError
        computed['soni'] = soni
    except (ValueError, TypeError):
        xatolar.append(f'Soni noto\'g\'ri: "{vals.get("soni")}"')

    # ── Majburiy: sana ──────────────────────────────────────────
    sana_str = vals.get('ish_sanasi', '').strip()
    if not sana_str:
        xatolar.append('ish_sanasi bo\'sh')
    else:
        try:
            sana = datetime.strptime(sana_str[:10], '%Y-%m-%d').date()
            computed['ish_sanasi'] = sana.isoformat()
        except ValueError:
            xatolar.append(f'Sana formati noto\'g\'ri: "{sana_str}" (YYYY-MM-DD kerak)')

    # ── Turi bo'yicha validatsiya ────────────────────────────────
    if turi in ('kroy', 'rezak'):
        _validate_kroy_row(vals, computed, xatolar, ogohlantirishlar)
    elif turi == 'zakatovka':
        _validate_zakatovka_row(vals, computed, xatolar, ogohlantirishlar)
    elif turi == 'kosib':
        _validate_kosib_row(vals, computed, xatolar, ogohlantirishlar)

    if xatolar:
        holat = 'xato'
    elif ogohlantirishlar:
        holat = 'ogohlantirish'
    else:
        holat = 'ok'

    return {
        'satr_num'         : satr_num,
        'holat'            : holat,
        'xatolar'          : xatolar,
        'ogohlantirishlar' : ogohlantirishlar,
        'vals'             : vals,
        'computed'         : computed,
    }


def _validate_kroy_row(vals, computed, xatolar, ogohlantirishlar):
    """Kroy / Rezak qatorini tekshirish"""
    teri_nomi = vals.get('teri_nomi', '').strip()
    if not teri_nomi:
        xatolar.append('teri_nomi bo\'sh')
        return

    teri_qs = Xomashyo.objects.filter(
        nomi__icontains=teri_nomi,
        category__name__iexact='teri',
        holati='active'
    )
    teri = teri_qs.first()
    if not teri:
        xatolar.append(f'Teri topilmadi: "{teri_nomi}"')
        return
    computed['teri_id'] = teri.id

    # Variant
    variant_rang = vals.get('teri_variant_rang', '').strip()
    if variant_rang:
        variant = XomashyoVariant.objects.filter(
            xomashyo=teri, rang__iexact=variant_rang
        ).first()
        if variant:
            computed['teri_variant_id'] = variant.id
        else:
            ogohlantirishlar.append(f'Teri variant topilmadi: "{variant_rang}" — variant ishlatilmaydi')

    # Sarfi
    mahsulot = m.Product.objects.filter(id=computed.get('mahsulot_id')).first()
    default_sarf = float(mahsulot.teri_sarfi) if mahsulot else 0

    sarfi_str = vals.get('teri_sarfi', '').strip()
    if sarfi_str:
        try:
            sarfi = float(sarfi_str)
            if sarfi <= 0:
                raise ValueError
            computed['teri_sarfi'] = sarfi
        except ValueError:
            ogohlantirishlar.append(f'teri_sarfi noto\'g\'ri: "{sarfi_str}" — default ({default_sarf}) ishlatiladi')
            computed['teri_sarfi'] = default_sarf
    else:
        if default_sarf > 0:
            ogohlantirishlar.append(f'teri_sarfi kiritilmagan — mahsulot default ({default_sarf} Dm) ishlatiladi')
            computed['teri_sarfi'] = default_sarf
        else:
            xatolar.append('teri_sarfi kiritilmagan va mahsulotda default sarfi yo\'q')

    # Miqdor yetarliligi
    soni = computed.get('soni', 0)
    sarfi = computed.get('teri_sarfi', 0)
    kerak = sarfi * soni

    variant_id = computed.get('teri_variant_id')
    if variant_id:
        v = XomashyoVariant.objects.filter(id=variant_id).first()
        mavjud = float(v.miqdori) if v else 0
    else:
        mavjud = float(teri.miqdori)

    if kerak > mavjud:
        xatolar.append(
            f'Teri yetarli emas — Kerak: {kerak:.2f}, Mavjud: {mavjud:.2f} ({teri_nomi})'
        )
    else:
        computed['teri_kerak'] = kerak

    # Astar (ixtiyoriy)
    astar_nomi = vals.get('astar_nomi', '').strip()
    if astar_nomi:
        astar = Xomashyo.objects.filter(
            nomi__icontains=astar_nomi,
            category__name__iexact='astar',
            holati='active'
        ).first()
        if astar:
            computed['astar_id'] = astar.id
            default_astar = float(
                m.Product.objects.filter(id=computed.get('mahsulot_id')).values_list('astar_sarfi', flat=True).first() or 0
            )
            sarfi_str2 = vals.get('astar_sarfi', '').strip()
            if sarfi_str2:
                try:
                    computed['astar_sarfi'] = float(sarfi_str2)
                except ValueError:
                    computed['astar_sarfi'] = default_astar
                    ogohlantirishlar.append(f'astar_sarfi noto\'g\'ri — default ({default_astar}) ishlatiladi')
            else:
                computed['astar_sarfi'] = default_astar
                if default_astar == 0:
                    ogohlantirishlar.append('astar_sarfi kiritilmagan va default 0')
        else:
            ogohlantirishlar.append(f'Astar topilmadi: "{astar_nomi}" — astar ishlatilmaydi')


def _validate_zakatovka_row(vals, computed, xatolar, ogohlantirishlar):
    """Zakatovka qatorini tekshirish"""
    mustaqil = vals.get('mustaqil_ish', 'yo\'q').lower().strip() in ('ha', 'yes', '1', 'true')
    computed['mustaqil_ish'] = mustaqil

    kroy_nomi = vals.get('kroy_xomashyo', '').strip()
    if kroy_nomi:
        kroy = Xomashyo.objects.filter(
            nomi__icontains=kroy_nomi,
            category__name__iexact='kroy',
            category__turi='process',
            holati='active'
        ).first()
        if kroy:
            computed['kroy_id'] = kroy.id
            soni = computed.get('soni', 0)
            if float(kroy.miqdori) < soni:
                xatolar.append(
                    f'Kroy yetarli emas — Kerak: {soni}, Mavjud: {kroy.miqdori} ({kroy_nomi})'
                )
        else:
            ogohlantirishlar.append(f'Kroy xomashyo topilmadi: "{kroy_nomi}"')
    elif not mustaqil:
        xatolar.append('Kroy xomashyo ko\'rsatilmagan. "mustaqil_ish" = "ha" qiling yoki kroy tanlang.')


def _validate_kosib_row(vals, computed, xatolar, ogohlantirishlar):
    """Kosib qatorini tekshirish"""
    mustaqil = vals.get('mustaqil_ish', 'yo\'q').lower().strip() in ('ha', 'yes', '1', 'true')
    computed['mustaqil_ish'] = mustaqil

    # Padoj — majburiy
    padoj_nomi = vals.get('padoj_nomi', '').strip()
    if not padoj_nomi:
        xatolar.append('padoj_nomi bo\'sh (kosib uchun majburiy)')
    else:
        padoj = Xomashyo.objects.filter(
            nomi__icontains=padoj_nomi,
            holati='active'
        ).first()
        if not padoj:
            xatolar.append(f'Padoj topilmadi: "{padoj_nomi}"')
        else:
            computed['padoj_id'] = padoj.id
            soni = computed.get('soni', 0)
            if float(padoj.miqdori) < soni:
                xatolar.append(
                    f'Padoj yetarli emas — Kerak: {soni}, Mavjud: {padoj.miqdori}'
                )
            # Padoj variant
            padoj_variant_rang = vals.get('padoj_variant_rang', '').strip()
            if padoj_variant_rang:
                pv = XomashyoVariant.objects.filter(
                    xomashyo=padoj, rang__iexact=padoj_variant_rang
                ).first()
                if pv:
                    computed['padoj_variant_id'] = pv.id
                else:
                    ogohlantirishlar.append(f'Padoj variant topilmadi: "{padoj_variant_rang}"')

    # Zakatovka (ixtiyoriy, lekin mustaqil_ish = 'yo\'q' bo'lsa kerak)
    zak_nomi = vals.get('zakatovka_nomi', '').strip()
    if zak_nomi:
        zak = Xomashyo.objects.filter(
            nomi__icontains=zak_nomi,
            category__name__iexact='zakatovka',
            category__turi='process',
            holati='active'
        ).first()
        if zak:
            computed['zakatovka_id'] = zak.id
            soni = computed.get('soni', 0)
            if float(zak.miqdori) < soni:
                xatolar.append(
                    f'Zakatovka yetarli emas — Kerak: {soni}, Mavjud: {zak.miqdori}'
                )
        else:
            ogohlantirishlar.append(f'Zakatovka topilmadi: "{zak_nomi}"')
    elif not mustaqil:
        xatolar.append('zakatovka_nomi ko\'rsatilmagan. "mustaqil_ish" = "ha" qiling yoki zakatovka tanlang.')

    # Variant
    computed['variant_rang']   = vals.get('variant_rang', '').strip()
    computed['variant_razmer'] = vals.get('variant_razmer', '').strip()


# ══════════════════════════════════════════════════════════════════
# SATR SAQLASH (mavjud IshQoshishView logikasiga mos)
# ══════════════════════════════════════════════════════════════════
def _satr_saqlash(satr, ishchi, turi, xomashyo_map):
    """Bitta tasdiqlangan satrni saqlash — IshQoshishView POST ga mos logika"""
    from xomashyo.models import IshXomashyo

    c = satr['computed']
    mahsulot = m.Product.objects.get(id=c['mahsulot_id'])
    soni     = int(c['soni'])
    sana     = datetime.strptime(c['ish_sanasi'], '%Y-%m-%d').date()

    ish = m.Ish.objects.create(
        ishchi  = ishchi,
        mahsulot= mahsulot,
        soni    = soni,
        status  = 'yangi',
        sana    = sana,
    )

    if turi in ('kroy', 'rezak'):
        _saqlash_kroy(ish, ishchi, mahsulot, soni, sana, c, xomashyo_map)

    elif turi == 'zakatovka':
        _saqlash_zakatovka(ish, mahsulot, soni, c)

    elif turi == 'kosib':
        _saqlash_kosib(ish, ishchi, mahsulot, soni, c)


def _get_or_create_jarayon(mahsulot, category_name, miqdor):
    """IshQoshishView._get_or_create_jarayon_xomashyo bilan bir xil"""
    category, _ = XomashyoCategory.objects.get_or_create(
        name=category_name, defaults={'turi': 'process'}
    )
    xom = Xomashyo.objects.filter(
        mahsulot=mahsulot, category=category, olchov_birligi='dona'
    ).first()
    if xom:
        xom.miqdori += Decimal(str(miqdor))
        xom.save(update_fields=['miqdori', 'updated_at'])
    else:
        xom = Xomashyo.objects.create(
            mahsulot=mahsulot, category=category,
            nomi=f"{mahsulot.nomi} - {category_name.title()}",
            miqdori=Decimal(str(miqdor)), olchov_birligi='dona',
            holati='active',
        )
    return xom


def _saqlash_kroy(ish, ishchi, mahsulot, soni, sana, c, xomashyo_map):
    from xomashyo.models import IshXomashyo

    # Teri sarfi (TeriSarfi + IshXomashyo)
    teri_id      = c.get('teri_id') or xomashyo_map.get(f"{mahsulot.nomi}:teri")
    teri         = Xomashyo.objects.get(id=teri_id)
    teri_variant = None
    if c.get('teri_variant_id'):
        teri_variant = XomashyoVariant.objects.get(id=c['teri_variant_id'])

    teri_sarfi = Decimal(str(c.get('teri_sarfi', mahsulot.teri_sarfi)))
    teri_jami  = teri_sarfi * soni

    m.TeriSarfi.objects.create(
        ish=ish, ishchi=ishchi, xomashyo=teri,
        variant=teri_variant, miqdor=teri_jami, sana=timezone.now()
    )
    IshXomashyo.objects.create(ish=ish, xomashyo=teri, variant=teri_variant, miqdor=teri_jami)

    # Astar (ixtiyoriy)
    astar_id = c.get('astar_id') or xomashyo_map.get(f"{mahsulot.nomi}:astar")
    if astar_id:
        astar       = Xomashyo.objects.get(id=astar_id)
        astar_sarfi = Decimal(str(c.get('astar_sarfi', mahsulot.astar_sarfi)))
        astar_jami  = astar_sarfi * soni
        if astar_jami > 0:
            IshXomashyo.objects.create(ish=ish, xomashyo=astar, miqdor=astar_jami)

    # Kroy jarayon
    kroy_xom = _get_or_create_jarayon(mahsulot, 'kroy', soni)
    IshXomashyo.objects.create(ish=ish, xomashyo=kroy_xom, miqdor=Decimal(str(soni)))


def _saqlash_zakatovka(ish, mahsulot, soni, c):
    from xomashyo.models import IshXomashyo

    mustaqil = c.get('mustaqil_ish', False)

    if not mustaqil and c.get('kroy_id'):
        kroy = Xomashyo.objects.get(id=c['kroy_id'])
        kroy.miqdori -= Decimal(str(soni))
        kroy.save(update_fields=['miqdori', 'updated_at'])
        IshXomashyo.objects.create(ish=ish, xomashyo=kroy, miqdor=Decimal(str(soni)))

    zak_xom = _get_or_create_jarayon(mahsulot, 'zakatovka', soni)
    IshXomashyo.objects.create(ish=ish, xomashyo=zak_xom, miqdor=Decimal(str(soni)))


def _saqlash_kosib(ish, ishchi, mahsulot, soni, c):
    from xomashyo.models import IshXomashyo

    mustaqil = c.get('mustaqil_ish', False)

    # Padoj
    padoj = Xomashyo.objects.get(id=c['padoj_id'])
    if c.get('padoj_variant_id'):
        pv = XomashyoVariant.objects.get(id=c['padoj_variant_id'])
        pv.miqdori -= Decimal(str(soni)); pv.save()
        IshXomashyo.objects.create(ish=ish, xomashyo=padoj, variant=pv, miqdor=Decimal(str(soni)))
    else:
        padoj.miqdori -= Decimal(str(soni))
        padoj.save(update_fields=['miqdori', 'updated_at'])
        IshXomashyo.objects.create(ish=ish, xomashyo=padoj, miqdor=Decimal(str(soni)))

    # Zakatovka (mustaqil emas bo'lsa)
    if not mustaqil and c.get('zakatovka_id'):
        zak = Xomashyo.objects.get(id=c['zakatovka_id'])
        zak.miqdori -= Decimal(str(soni)); zak.save(update_fields=['miqdori', 'updated_at'])
        IshXomashyo.objects.create(ish=ish, xomashyo=zak, miqdor=Decimal(str(soni)))

    # Mahsulot variant
    rang   = c.get('variant_rang', '')
    razmer = c.get('variant_razmer', '')
    variant, created = m.ProductVariant.objects.get_or_create(
        product=mahsulot, rang=rang, razmer=razmer,
        defaults={'stock': soni, 'price': mahsulot.narxi}
    )
    if not created:
        variant.stock += soni; variant.save()

    # Kosib jarayon
    kosib_xom = _get_or_create_jarayon(mahsulot, 'kosib', soni)
    IshXomashyo.objects.create(ish=ish, xomashyo=kosib_xom, miqdor=Decimal(str(soni)))
    mahsulot.refresh_from_db()