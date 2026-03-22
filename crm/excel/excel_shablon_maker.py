# excel_shablon_maker.py
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, Protection
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import io

# ─── Ranglar ──────────────────────────────────────────────────────
COLORS = {
    'header_bg'   : 'FF2D3748',   # qoraga yaqin ko'k
    'header_fg'   : 'FFFFFFFF',
    'req_bg'      : 'FFE8F4FD',   # majburiy ustun — och ko'k
    'opt_bg'      : 'FFFFF9E6',   # ixtiyoriy  — och sariq
    'info_bg'     : 'FFF0FFF4',   # ma'lumot    — och yashil
    'border_color': 'FFB0BEC5',
    'example_fg'  : 'FF546E7A',
    'title_bg'    : 'FF1A365D',
    'title_fg'    : 'FFFFFFFF',
}

thin  = Side(style='thin',   color=COLORS['border_color'])
thick = Side(style='medium', color='FF2D3748')
border_all   = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
border_thick = Border(left=thick, right=thick, top=thick, bottom=thick)

def hdr_font(bold=True):
    return Font(name='Arial', bold=bold, color=COLORS['header_fg'], size=10)

def body_font(bold=False, color='FF1A202C'):
    return Font(name='Arial', bold=bold, color=color, size=10)

def fill(hex_color):
    return PatternFill('solid', start_color=hex_color, fgColor=hex_color)

def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def left():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)

def style_header_row(ws, row_num, cols_count):
    for c in range(1, cols_count + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.font      = hdr_font()
        cell.fill      = fill(COLORS['header_bg'])
        cell.alignment = center()
        cell.border    = border_all

def style_data_row(ws, row_num, cols_meta):
    """cols_meta: list of ('required'|'optional'|'info')"""
    for c, kind in enumerate(cols_meta, 1):
        cell = ws.cell(row=row_num, column=c)
        bg   = COLORS['req_bg'] if kind == 'required' else \
               COLORS['opt_bg'] if kind == 'optional' else COLORS['info_bg']
        cell.fill      = fill(bg)
        cell.alignment = left()
        cell.border    = border_all
        cell.font      = body_font()

def add_title(ws, title, subtitle, cols_count):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols_count)
    t = ws.cell(1, 1, title)
    t.font      = Font(name='Arial', bold=True, size=13, color=COLORS['title_fg'])
    t.fill      = fill(COLORS['title_bg'])
    t.alignment = center()

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols_count)
    s = ws.cell(2, 1, subtitle)
    s.font      = Font(name='Arial', size=9, color='FF546E7A', italic=True)
    s.fill      = fill('FFF7FAFC')
    s.alignment = center()

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18

def add_legend(ws, row, cols_count):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols_count)
    cell = ws.cell(row, 1,
        "🔵 Ko'k = majburiy  |  🟡 Sariq = ixtiyoriy  |  "
        "Sana formati: YYYY-MM-DD  |  Ha/Yo'q: ha yoki yo'q"
    )
    cell.font      = Font(name='Arial', size=8, color='FF546E7A', italic=True)
    cell.fill      = fill('FFFAFAFA')
    cell.alignment = left()
    cell.border    = border_all

def freeze_and_filter(ws, header_row):
    ws.freeze_panes = ws.cell(header_row + 1, 1)
    ws.auto_filter.ref = ws.dimensions

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def add_dropdown(ws, col, first_row, last_row, formula):
    dv = DataValidation(type='list', formula1=formula, allow_blank=True)
    dv.sqref = f"{get_column_letter(col)}{first_row}:{get_column_letter(col)}{last_row}"
    ws.add_data_validation(dv)

# ══════════════════════════════════════════════════════════════════
# SHABLON GENERATORLAR
# ══════════════════════════════════════════════════════════════════

def make_kroy_shablon(ishchi_ism='', mahsulotlar=None, terilar=None, astarlar=None):
    """Kroy / Rezak shablon"""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Ishlar'

    # Ma'lumot varaq — dropdown uchun
    ref_ws = wb.create_sheet('Ma\'lumotlar')
    _fill_ref_sheet(ref_ws, mahsulotlar or [], terilar or [], astarlar or [], [])

    COLS = [
        ('mahsulot_nomi',     'Mahsulot nomi *',          'required', 25),
        ('soni',              'Soni *',                   'required', 10),
        ('ish_sanasi',        'Ish sanasi * (YYYY-MM-DD)', 'required', 18),
        ('teri_nomi',         'Teri nomi *',               'required', 22),
        ('teri_variant_rang', 'Teri variant rang',         'optional', 18),
        ('teri_sarfi',        'Teri sarfi (Dm) bitta',     'optional', 18),
        ('astar_nomi',        'Astar nomi',                'optional', 22),
        ('astar_variant_rang','Astar variant rang',        'optional', 18),
        ('astar_sarfi',       'Astar sarfi (Dm) bitta',    'optional', 18),
        ('izoh',              'Izoh',                      'optional', 20),
    ]

    n_cols = len(COLS)
    add_title(ws,
        f'Kroy / Rezak — Ish jadvali ({ishchi_ism})',
        'Majburiy ustunlar (*) to\'ldirilishi shart. Bir satr = bir ish yozuvi.',
        n_cols)
    add_legend(ws, 3, n_cols)

    # Header
    for c, (key, label, kind, w) in enumerate(COLS, 1):
        ws.cell(4, c, label)
    style_header_row(ws, 4, n_cols)
    ws.row_dimensions[4].height = 36

    # Namuna satr
    example = [
        'Sumka A', '5', '2025-03-10',
        'Qo\'ng\'ir teri', 'Jigarrang', '1.50',
        'Oq astar', '', '0.80', ''
    ]
    for c, val in enumerate(example, 1):
        ws.cell(5, c, val)
        ws.cell(5, c).font = body_font(color=COLORS['example_fg'])
        ws.cell(5, c).fill = fill('FFF7FAFC')
        ws.cell(5, c).border = border_all
        ws.cell(5, c).alignment = left()

    # Bo'sh ma'lumot satrlari (6–105)
    kinds = [x[2] for x in COLS]
    for row in range(6, 106):
        style_data_row(ws, row, kinds)

    # Dropdownlar — mahsulot (col 1), teri (col 4), astar (col 7)
    if mahsulotlar:
        add_dropdown(ws, 1, 6, 105, f"Ma\'lumotlar!$A$2:$A${len(mahsulotlar)+1}")
    if terilar:
        add_dropdown(ws, 4, 6, 105, f"Ma\'lumotlar!$B$2:$B${len(terilar)+1}")
    if astarlar:
        add_dropdown(ws, 7, 6, 105, f"Ma\'lumotlar!$C$2:$C${len(astarlar)+1}")

    # Ha/yo'q dropdown — ixtiyoriy emas lekin kerak bo'lsa
    set_col_widths(ws, [x[3] for x in COLS])
    freeze_and_filter(ws, 4)
    ws.sheet_view.showGridLines = False

    return wb


def make_zakatovka_shablon(ishchi_ism='', mahsulotlar=None, kroy_xomashyolar=None):
    """Zakatovka shablon"""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Ishlar'

    ref_ws = wb.create_sheet('Ma\'lumotlar')
    _fill_ref_sheet(ref_ws, mahsulotlar or [], [], [], kroy_xomashyolar or [])

    COLS = [
        ('mahsulot_nomi',     'Mahsulot nomi *',          'required', 25),
        ('soni',              'Soni *',                   'required', 10),
        ('ish_sanasi',        'Ish sanasi * (YYYY-MM-DD)', 'required', 18),
        ('kroy_xomashyo',     'Kroy xomashyo',             'optional', 25),
        ('mustaqil_ish',      'Mustaqil ish (ha/yo\'q)',   'optional', 18),
        ('izoh',              'Izoh',                      'optional', 20),
    ]

    n_cols = len(COLS)
    add_title(ws,
        f'Zakatovka — Ish jadvali ({ishchi_ism})',
        'kroy_xomashyo bo\'lmasa "mustaqil_ish" ustuniga "ha" yozing.',
        n_cols)
    add_legend(ws, 3, n_cols)

    for c, (key, label, kind, w) in enumerate(COLS, 1):
        ws.cell(4, c, label)
    style_header_row(ws, 4, n_cols)
    ws.row_dimensions[4].height = 36

    example = ['Sumka A', '5', '2025-03-10', 'Sumka A - kroy', 'yo\'q', '']
    for c, val in enumerate(example, 1):
        ws.cell(5, c, val)
        ws.cell(5, c).font = body_font(color=COLORS['example_fg'])
        ws.cell(5, c).fill = fill('FFF7FAFC')
        ws.cell(5, c).border = border_all
        ws.cell(5, c).alignment = left()

    kinds = [x[2] for x in COLS]
    for row in range(6, 106):
        style_data_row(ws, row, kinds)

    if mahsulotlar:
        add_dropdown(ws, 1, 6, 105, f"Ma\'lumotlar!$A$2:$A${len(mahsulotlar)+1}")
    if kroy_xomashyolar:
        add_dropdown(ws, 4, 6, 105, f"Ma\'lumotlar!$D$2:$D${len(kroy_xomashyolar)+1}")
    add_dropdown(ws, 5, 6, 105, '"ha,yo\'q"')

    set_col_widths(ws, [x[3] for x in COLS])
    freeze_and_filter(ws, 4)
    ws.sheet_view.showGridLines = False
    return wb


def make_kosib_shablon(ishchi_ism='', mahsulotlar=None, padojlar=None, zakatovkalar=None):
    """Kosib shablon"""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Ishlar'

    ref_ws = wb.create_sheet('Ma\'lumotlar')
    _fill_ref_sheet(ref_ws, mahsulotlar or [], [], padojlar or [], zakatovkalar or [])

    COLS = [
        ('mahsulot_nomi',      'Mahsulot nomi *',              'required', 25),
        ('soni',               'Soni *',                      'required', 10),
        ('ish_sanasi',         'Ish sanasi * (YYYY-MM-DD)',    'required', 18),
        ('padoj_nomi',         'Padoj nomi *',                 'required', 22),
        ('padoj_variant_rang', 'Padoj variant rang',           'optional', 18),
        ('variant_rang',       'Mahsulot variant rang',        'optional', 20),
        ('variant_razmer',     'Mahsulot variant razmer',      'optional', 20),
        ('zakatovka_nomi',     'Zakatovka xomashyo',           'optional', 25),
        ('mustaqil_ish',       'Mustaqil ish (ha/yo\'q)',      'optional', 18),
        ('izoh',               'Izoh',                         'optional', 20),
    ]

    n_cols = len(COLS)
    add_title(ws,
        f'Kosib — Ish jadvali ({ishchi_ism})',
        'padoj_nomi majburiy. zakatovka_nomi bo\'lmasa mustaqil_ish = "ha" yozing.',
        n_cols)
    add_legend(ws, 3, n_cols)

    for c, (key, label, kind, w) in enumerate(COLS, 1):
        ws.cell(4, c, label)
    style_header_row(ws, 4, n_cols)
    ws.row_dimensions[4].height = 36

    example = ['Sumka A', '5', '2025-03-10', 'Padoj B', 'Qora', 'Qo\'ng\'ir', '40', 'Sumka A - zakatovka', 'yo\'q', '']
    for c, val in enumerate(example, 1):
        ws.cell(5, c, val)
        ws.cell(5, c).font = body_font(color=COLORS['example_fg'])
        ws.cell(5, c).fill = fill('FFF7FAFC')
        ws.cell(5, c).border = border_all
        ws.cell(5, c).alignment = left()

    kinds = [x[2] for x in COLS]
    for row in range(6, 106):
        style_data_row(ws, row, kinds)

    if mahsulotlar:
        add_dropdown(ws, 1, 6, 105, f"Ma\'lumotlar!$A$2:$A${len(mahsulotlar)+1}")
    if padojlar:
        add_dropdown(ws, 4, 6, 105, f"Ma\'lumotlar!$C$2:$C${len(padojlar)+1}")
    if zakatovkalar:
        add_dropdown(ws, 8, 6, 105, f"Ma\'lumotlar!$D$2:$D${len(zakatovkalar)+1}")
    add_dropdown(ws, 9, 6, 105, '"ha,yo\'q"')

    set_col_widths(ws, [x[3] for x in COLS])
    freeze_and_filter(ws, 4)
    ws.sheet_view.showGridLines = False
    return wb


def make_pardoz_shablon(ishchi_ism='', mahsulotlar=None):
    """Pardoz / boshqa ishchilar uchun oddiy shablon"""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Ishlar'

    ref_ws = wb.create_sheet('Ma\'lumotlar')
    _fill_ref_sheet(ref_ws, mahsulotlar or [], [], [], [])

    COLS = [
        ('mahsulot_nomi', 'Mahsulot nomi *',          'required', 25),
        ('soni',          'Soni *',                   'required', 10),
        ('ish_sanasi',    'Ish sanasi * (YYYY-MM-DD)', 'required', 18),
        ('izoh',          'Izoh',                      'optional', 30),
    ]

    n_cols = len(COLS)
    add_title(ws, f'Pardoz — Ish jadvali ({ishchi_ism})', 'Barcha majburiy ustunlarni to\'ldiring.', n_cols)
    add_legend(ws, 3, n_cols)

    for c, (key, label, kind, w) in enumerate(COLS, 1):
        ws.cell(4, c, label)
    style_header_row(ws, 4, n_cols)
    ws.row_dimensions[4].height = 36

    example = ['Sumka A', '5', '2025-03-10', '']
    for c, val in enumerate(example, 1):
        ws.cell(5, c, val)
        ws.cell(5, c).font = body_font(color=COLORS['example_fg'])
        ws.cell(5, c).fill = fill('FFF7FAFC')
        ws.cell(5, c).border = border_all
        ws.cell(5, c).alignment = left()

    kinds = [x[2] for x in COLS]
    for row in range(6, 106):
        style_data_row(ws, row, kinds)

    if mahsulotlar:
        add_dropdown(ws, 1, 6, 105, f"Ma\'lumotlar!$A$2:$A${len(mahsulotlar)+1}")

    set_col_widths(ws, [x[3] for x in COLS])
    freeze_and_filter(ws, 4)
    ws.sheet_view.showGridLines = False
    return wb


def _fill_ref_sheet(ws, mahsulotlar, terilar, padojlar, kroy_zak):
    """Ma'lumotlar varaqini to'ldirish — dropdown source"""
    headers = ['Mahsulotlar', 'Terilar', 'Padojlar/Astarlar', 'Kroy/Zakatovka']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(name='Arial', bold=True, size=9, color='FFFFFFFF')
        cell.fill = fill('FF4A5568')
        cell.alignment = center()
        ws.column_dimensions[get_column_letter(c)].width = 28

    lists = [mahsulotlar, terilar, padojlar, kroy_zak]
    for col, lst in enumerate(lists, 1):
        for row, val in enumerate(lst, 2):
            ws.cell(row, col, val).font = Font(name='Arial', size=9)


def get_shablon_bytes(turi, **kwargs):
    """Shablon wb → bytes (HTTP response uchun)"""
    makers = {
        'kroy'      : make_kroy_shablon,
        'rezak'     : make_kroy_shablon,
        'kosib'     : make_kosib_shablon,
        'zakatovka' : make_zakatovka_shablon,
        'pardoz'    : make_pardoz_shablon,
        'pardozchi' : make_pardoz_shablon,
    }
    maker = makers.get(turi.lower(), make_pardoz_shablon)
    wb = maker(**kwargs)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Test ──────────────────────────────────────────────────────────
if __name__ == '__main__':
    mahsulotlar = ['Sumka A', 'Sumka B', 'Sumka C', 'Sumka D']
    terilar     = ["Qo'ng'ir teri #1", "Qo'ng'ir teri #2", 'Qora teri', 'Jigarrang teri']
    astarlar    = ['Oq astar', "Ko'k astar", 'Sariq astar']
    padojlar    = ['Padoj A', 'Padoj B', 'Padoj C']
    kroy_zak    = ['Sumka A - kroy', 'Sumka B - kroy']

    for turi, kwargs in [
        ('kroy',      dict(ishchi_ism='Alisher K.', mahsulotlar=mahsulotlar, terilar=terilar, astarlar=astarlar)),
        ('zakatovka', dict(ishchi_ism='Dilnoza Y.', mahsulotlar=mahsulotlar, kroy_xomashyolar=kroy_zak)),
        ('kosib',     dict(ishchi_ism='Bahodir T.', mahsulotlar=mahsulotlar, padojlar=padojlar, zakatovkalar=kroy_zak)),
        ('pardoz',    dict(ishchi_ism='Nodira S.',  mahsulotlar=mahsulotlar)),
    ]:
        data = get_shablon_bytes(turi, **kwargs)
        path = f'/home/claude/excel_ish/shablon_{turi}.xlsx'
        with open(path, 'wb') as f:
            f.write(data)
        print(f'✓ {path}  ({len(data):,} bytes)')